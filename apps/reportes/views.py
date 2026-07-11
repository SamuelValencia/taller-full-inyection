from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

from apps.decorators import admin_o_recepcionista_requerido, rol_requerido
from apps.ordenes.models import OrdenTrabajo, DetalleOrdenTrabajo
from apps.clientes.models import Cliente
from apps.vehiculos.models import Vehiculo
from apps.mantenimiento.models import HistorialMantenimiento
from apps.inventario.models import Repuesto


def render_pdf(template_path, context):
    template = get_template(template_path)
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("utf-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type="application/pdf")
    return HttpResponse("Error generando PDF", status=500)


@admin_o_recepcionista_requerido
def index(request):
    total_ordenes = OrdenTrabajo.objects.count()
    total_clientes = Cliente.objects.filter(activo=True).count()
    total_vehiculos = Vehiculo.objects.filter(activo=True).count()
    ordenes_por_estado = {
        "recibidas": OrdenTrabajo.objects.filter(estado="RECIBIDA").count(),
        "en_proceso": OrdenTrabajo.objects.filter(estado="EN_PROCESO").count(),
        "finalizadas": OrdenTrabajo.objects.filter(estado="FINALIZADA").count(),
        "entregadas": OrdenTrabajo.objects.filter(estado="ENTREGADA").count(),
    }
    return render(request, "reportes/index.html", {
        "total_ordenes": total_ordenes,
        "total_clientes": total_clientes,
        "total_vehiculos": total_vehiculos,
        "ordenes_por_estado": ordenes_por_estado,
    })


@admin_o_recepcionista_requerido
def orden_pdf(request, pk):
    orden = get_object_or_404(
        OrdenTrabajo.objects.select_related("vehiculo", "cliente", "tecnico_asignado", "recepcionista"),
        pk=pk
    )
    detalles = orden.detalles.all()
    context = {
        "orden": orden,
        "detalles": detalles,
        "taller_nombre": settings.TALLER_NOMBRE,
        "taller_telefono": settings.TALLER_TELEFONO,
        "taller_direccion": settings.TALLER_DIRECCION,
    }
    return render_pdf("reportes/orden_pdf.html", context)


@admin_o_recepcionista_requerido
def historial_vehiculo_pdf(request, pk):
    vehiculo = get_object_or_404(Vehiculo.objects.select_related("cliente"), pk=pk)
    historial = HistorialMantenimiento.objects.filter(vehiculo=vehiculo).order_by("-fecha_servicio")
    context = {
        "vehiculo": vehiculo,
        "historial": historial,
        "taller_nombre": settings.TALLER_NOMBRE,
        "taller_telefono": settings.TALLER_TELEFONO,
        "taller_direccion": settings.TALLER_DIRECCION,
    }
    return render_pdf("reportes/historial_pdf.html", context)


# Exportaciones — solo ADMIN y GERENTE
_solo_admin_gerente = rol_requerido('ADMIN', 'GERENTE')


@_solo_admin_gerente
def exportar_clientes_excel(request):
    clientes = Cliente.objects.filter(activo=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    headers = ["Documento", "Nombre", "Telefono", "Email", "Direccion", "Activo"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for cliente in clientes:
        ws.append([
            cliente.tipo_documento,
            cliente.nombre_completo,
            cliente.telefono or "",
            cliente.email or "",
            cliente.direccion or "",
            "Si" if cliente.activo else "No"
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
    wb.save(response)
    return response


@_solo_admin_gerente
def exportar_vehiculos_excel(request):
    vehiculos = Vehiculo.objects.filter(activo=True).select_related("cliente")
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehiculos"
    headers = ["Placa", "Marca", "Modelo", "Anio", "Color", "Cliente", "Activo"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for vehiculo in vehiculos:
        ws.append([
            vehiculo.placa, vehiculo.marca, vehiculo.modelo,
            vehiculo.anio, vehiculo.color,
            vehiculo.cliente.nombre_completo,
            "Si" if vehiculo.activo else "No"
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="vehiculos.xlsx"'
    wb.save(response)
    return response


@_solo_admin_gerente
def exportar_ordenes_excel(request):
    ordenes = OrdenTrabajo.objects.select_related("cliente", "vehiculo", "tecnico_asignado")
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordenes"
    headers = ["N Orden", "Cliente", "Vehiculo", "Estado", "Prioridad", "Tecnico", "Fecha Ingreso", "Total"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for orden in ordenes:
        ws.append([
            orden.numero_orden,
            orden.cliente.nombre_completo,
            orden.vehiculo.placa,
            orden.get_estado_display(),
            orden.get_prioridad_display(),
            orden.tecnico_asignado.get_full_name() if orden.tecnico_asignado else "",
            orden.fecha_ingreso.strftime("%d/%m/%Y %H:%M"),
            float(orden.costo_total)
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ordenes.xlsx"'
    wb.save(response)
    return response


@_solo_admin_gerente
def exportar_inventario_excel(request):
    repuestos = Repuesto.objects.filter(activo=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    headers = ["Codigo", "Nombre", "Categoria", "Marca", "Stock Actual", "Stock Minimo", "Precio Compra", "Precio Venta", "Valor Inventario"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for repuesto in repuestos:
        ws.append([
            repuesto.codigo, repuesto.nombre,
            repuesto.categoria.nombre if repuesto.categoria else "",
            repuesto.marca or "",
            repuesto.stock_actual, repuesto.stock_minimo,
            float(repuesto.precio_compra), float(repuesto.precio_venta),
            float(repuesto.valor_inventario)
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario.xlsx"'
    wb.save(response)
    return response


# Reportes gerenciales — solo ADMIN y GERENTE
@_solo_admin_gerente
def productividad_taller(request):
    from django.db.models import Count, Sum, Avg
    from django.db.models.functions import TruncDate
    from datetime import datetime, timedelta

    fecha_inicio = datetime.now().date() - timedelta(days=30)
    ordenes_completadas = OrdenTrabajo.objects.filter(
        fecha_entrega_real__gte=fecha_inicio,
        estado__in=['FINALIZADA', 'ENTREGADA']
    )
    total_ordenes = ordenes_completadas.count()
    total_ingresos = ordenes_completadas.aggregate(total=Sum('costo_total'))['total'] or 0
    promedio_por_orden = ordenes_completadas.aggregate(promedio=Avg('costo_total'))['promedio'] or 0
    ordenes_por_dia = ordenes_completadas.annotate(
        fecha=TruncDate('fecha_entrega_real')
    ).values('fecha').annotate(count=Count('id')).order_by('fecha')

    from apps.usuarios.models import Usuario
    tecnicos = Usuario.objects.filter(
        rol__codigo="MECANICO",
        ordenes_trabajo_asignadas__fecha_entrega_real__gte=fecha_inicio,
        ordenes_trabajo_asignadas__estado__in=['FINALIZADA', 'ENTREGADA']
    ).annotate(ordenes_completadas=Count('ordenes_trabajo_asignadas')).order_by('-ordenes_completadas')[:5]

    context = {
        "fecha_inicio": fecha_inicio,
        "total_ordenes": total_ordenes,
        "total_ingresos": total_ingresos,
        "promedio_por_orden": promedio_por_orden,
        "ordenes_por_dia": ordenes_por_dia,
        "tecnicos": tecnicos,
    }
    return render(request, "reportes/productividad.html", context)


@admin_o_recepcionista_requerido
def servicios_realizados(request):
    from django.db.models import Count, Sum
    servicios = DetalleOrdenTrabajo.objects.filter(
        tipo='SERVICIO'
    ).values('descripcion').annotate(count=Count('id')).order_by('-count')[:20]
    repuestos = DetalleOrdenTrabajo.objects.filter(
        tipo='REPUESTO'
    ).values('descripcion').annotate(count=Count('id'), total_cantidad=Sum('cantidad')).order_by('-count')[:20]
    context = {"servicios": servicios, "repuestos": repuestos}
    return render(request, "reportes/servicios.html", context)


@admin_o_recepcionista_requerido
def reporte_mantenimientos(request):
    from apps.mantenimiento.models import ProgramaMantenimiento, HistorialMantenimiento
    from django.db.models import Count
    vencidos = ProgramaMantenimiento.objects.filter(estado=ProgramaMantenimiento.EstadoPrograma.VENCIDO).count()
    proximos = ProgramaMantenimiento.objects.filter(estado=ProgramaMantenimiento.EstadoPrograma.PROXIMO).count()
    al_dia = ProgramaMantenimiento.objects.filter(estado=ProgramaMantenimiento.EstadoPrograma.AL_DIA).count()
    tipos_comunes = HistorialMantenimiento.objects.values(
        'tipo_servicio__nombre'
    ).annotate(count=Count('id')).order_by('-count')[:10]
    context = {
        "vencidos": vencidos,
        "proximos": proximos,
        "al_dia": al_dia,
        "tipos_comunes": tipos_comunes,
    }
    return render(request, "reportes/mantenimientos.html", context)
